
import numpy as np
import pandas as pd
import requests
import ta
import tvDatafeed
import yfinance as yf
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from ta.wrapper import EMAIndicator
from tradingview_ta import TA_Handler, Interval
import time

#https://gcdn.bionluk.com/uploads/message/10c60a14-ca6d-43e9-9795-fc0d53afe262.txt   smi
#https://gcdn.bionluk.com/uploads/message/649c0c0a-ba70-4fc4-8085-2304fffcc5de.txt   rsi
#https://gcdn.bionluk.com/uploads/message/b3ab8bbe-7d1c-4255-8bb8-80be67169d39.txt   macd dema
#pyinstaller --hidden-import openpyxl application.py
#pyinstaller --onefile application.py
#479
'''
born = input("bist için 0, nasdaq için 1 giriniz:")
periyot = input("periyot giriniz:")
verisayisi = input("max veri için 0 yazınız, limitli ise veri sayısı giriniz:")
'''
born = "0"
periyot = "100"
verisayisi = "0"
def get_sma(prices, rate):
    return prices.rolling(rate).mean()

def get_bollinger_bands(prices , rate = 20 ):
    sma = get_sma(prices, rate)
    std = prices.rolling(rate).std()
    bollinger_up = sma + std * 2
    bollinger_down = sma - std * 2
    return bollinger_up, bollinger_down, sma

def MACD(ana,data, period_long=26, period_short=12, period_signal=9, column='Adj Close'):

    # short hesaplayalım
    short = data.ewm(span=period_short,adjust=False).mean()
    long = data.ewm(span=period_long,adjust=False).mean()
    ShortEMA = 2 * short - short.ewm(span=period_short,adjust=False).mean()
    # long
    LongEMA = 2 * long - long.ewm(span=period_long,adjust=False).mean()
    # MACD
    ana["MACD"] = ShortEMA - LongEMA
    # Signal
    signal = ana["MACD"].ewm(span=period_signal, adjust=False).mean()
    ana["Signal_Line"] = 2 * signal - signal.ewm(span=period_signal,adjust=False).mean()

    return data

def divideAdder(dividender, data):
    for i in range(data["close"].values.size):
        for j in range(len(dividender.dividends)):
            curr_date = dividender.dividends.index[j]
            if curr_date >= data.index[i]:
                divid = dividender.dividends[j]
                #data["open"].iat[i] -= divid
                #data["high"].iat[i] -= divid
                #data["low"].iat[i] -= divid
                #data["close"].iat[i] -= divid

def addvalue(sheet,value,indicator,row,time,where):
    sheet.cell(row,indicator.start+time*indicator.subber+where).value=value

class indicator:
    def __init__(self, name, sub, subber,color):
        self.name = name
        self.sub = sub
        self.subber = subber
        self.subberlist=[]
        self.color=color
        self.start=0
    def subbers(self,subberlist):
        self.subberlist=subberlist

class times:
    def __init__(self, name, sublist,color):
        self.name = name
        self.sublist = sublist
        self.color=color
        self.start=0

tv = tvDatafeed.TvDatafeed()
exchange = "BIST"
if born=="1":
    exchange= "NASDAQ"
    NASDAQ = "https://eoddata.com/stocklist/NASDAQ/A.htm"
    page = requests.get(NASDAQ)
    soup = BeautifulSoup(page.content, "html.parser")
    results = soup.findAll("tr", {"class": ["re","ro"]})
    results = str(results).split('</a>')
    BISTLIST = []
    for i in range(0, len(results)-2, 3):
        print(i)
        print(results[i].split('Chart for NASDAQ,')[1].split('">')[1])
        adder = TA_Handler(
            symbol=results[i].split('Chart for NASDAQ,')[1].split('">')[1],
            screener="america",
            exchange="NASDAQ",
            interval=Interval.INTERVAL_1_DAY
        )
        BISTLIST.append(adder)
else:
    BIST = "https://finans.mynet.com/borsa/hisseler/"
    page = requests.get(BIST)
    soup = BeautifulSoup(page.content, "html.parser")
    results = soup.findAll("strong", {"class": "mr-4"})
    results=str(results).split('">')
    BISTLIST = []
    for i in range(2,len(results),3):
        adder = TA_Handler(
            symbol=results[i].split(" ")[0],
            screener="turkey",
            exchange="BIST",
            interval=Interval.INTERVAL_1_DAY
        )
        BISTLIST.append(adder)

wb = Workbook()
sheet = wb.active
timelist=[]
indicatorlist=["smi","smi[1]","fark","ema5","ema5[1]","ema8","ema8[1]","ema13","ema13[1]","ema5_cross_ema8","ema8_cross_ema13","ema5_cross_ema13","smisignal","smi_smisignal","bolupsmi","basissmi","basis-smi","bollingerdownsmi","cross_basissmidif","bolupbasissmiper","cross_bolupsmi","bolupsmidif","bolupsmiper","boldownsmidif","boldownsmiper","cross_boldownsmi",
               "rsi","rsi[1]","farkrsi","boluprsi","basisrsi","basis-rsi","bollingerdownrsi","cross_basisrsidif","bolupbasisrsiper","cross_boluprsi","boluprsidif","boluprsiper","boldownrsidif","boldownrsiper","cross_boldownrsi",
               "cci","cci[1]","farkcci","lignemacdzerolag","lignesignal","kesişim_macd","k","d","kesişim_storsi","mom","mom[1]","farkmom","obv","obv[1]","farkobv","signal","kesişim_obv","tke","ematke","tkeemacross","tke[1]","tkefark","volume","close",
               "aroonup","aroonup[1]","aroondown","aroondown[1]","kesişim_aroon"]
ttimelist=["15min","30mins","1hour","4hours","1day","1week","1month"]
cellnumber=2
color=0
for i in indicatorlist:
    sheet.merge_cells(start_row=1, end_row=1, start_column=cellnumber, end_column=cellnumber + 6)
    sheet.cell(1,cellnumber).value=i
    yellow = "FFFF00"
    pink="FFC0CB"
    blue="B0E0E6"
    if color%2==1:
        sheet.cell(1, cellnumber).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
    else:
        sheet.cell(1, cellnumber).fill = PatternFill(start_color=pink, end_color=pink, fill_type="solid")
    for j,k in zip(ttimelist,range(7)):
        sheet.cell(2,cellnumber+k).value=i+"_"+j
        sheet.cell(2, cellnumber + k).fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
    cellnumber+=7
    color+=1
row=3
if(int(verisayisi) == 0):
    verisayisi=str(len(BISTLIST))
'''
sheet.cell(1, 387+42).value = "EMA10"
sheet.cell(1, 388+42).value = "EMA20"
sheet.cell(1, 389+42).value = "EMA50"
sheet.cell(1, 390+42).value = "EMA100"
sheet.cell(1, 391+42).value = "EMA200"
sheet.cell(1, 392+42).value = "EMA5"
sheet.cell(1, 393+42).value = "EMA5[1]"
sheet.cell(1, 394+42).value = "EMA8"
sheet.cell(1, 395+42).value = "EMA8[1]"
sheet.cell(1, 396+42).value = "EMA13"
sheet.cell(1, 397+42).value = "EMA13[1]"
sheet.cell(1, 398+42).value = "EMA5crossEMA8"
sheet.cell(1, 399+42).value = "EMA8crossEMA13"
sheet.cell(1, 400+42).value = "EMA5crossEMA13"
green="7FFF00"
sheet.cell(1, 387+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
sheet.cell(1, 388+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
sheet.cell(1, 389+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
sheet.cell(1, 390+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
sheet.cell(1, 391+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
sheet.cell(1, 392+42).fill = PatternFill(start_color=green, end_color=green, fill_type="solid") '''
for i in range(0,int(verisayisi)):
    timer=0
    print(i)
    tick=BISTLIST[i]
    print(tick.symbol)
    dividender = yf.Ticker(tick.symbol+".IS")
    sheet.cell(row,1).value=tick.symbol
    sheet.cell(row,1).hyperlink="https://tr.tradingview.com/chart/?symbol=BIST%3A"+tick.symbol
    for k,kk in zip([Interval.INTERVAL_15_MINUTES,Interval.INTERVAL_30_MINUTES,Interval.INTERVAL_1_HOUR,Interval.INTERVAL_4_HOURS,Interval.INTERVAL_1_DAY,Interval.INTERVAL_1_WEEK,Interval.INTERVAL_1_MONTH],[0,1,2,3,4,5,6]):
        print("---------")
        print("start: " + str(time.time()))
        tick.interval = k
        gg=int(periyot)
        interval=tvDatafeed.Interval.in_1_hour
        if k == Interval.INTERVAL_15_MINUTES:
            interval=tvDatafeed.Interval.in_15_minute
        elif k == Interval.INTERVAL_30_MINUTES:
            interval = tvDatafeed.Interval.in_30_minute
        elif k == Interval.INTERVAL_1_HOUR:
            interval = tvDatafeed.Interval.in_1_hour
        elif k == Interval.INTERVAL_4_HOURS:
            interval = tvDatafeed.Interval.in_4_hour
        elif k == Interval.INTERVAL_1_DAY:
            interval = tvDatafeed.Interval.in_daily
        elif k == Interval.INTERVAL_1_WEEK:
            interval = tvDatafeed.Interval.in_weekly
        elif k == Interval.INTERVAL_1_MONTH:
            interval = tvDatafeed.Interval.in_monthly
            gg=int(periyot)
        print("hist1: " + str(time.time()))
        data=tv.get_hist(tick.symbol,exchange=exchange,interval=interval,n_bars=gg)
        print("hist2: " + str(time.time()))
        try:
            data["open"] = pd.to_numeric(data["open"])
            data["high"] = pd.to_numeric(data["high"])
            data["low"] = pd.to_numeric(data["low"])
            data["close"] = pd.to_numeric(data["close"])
            data["volume"] = pd.to_numeric(data["volume"])
        except:
            break
        data.dropna(inplace=True)
        '''if(k == Interval.INTERVAL_1_DAY):
            sizzze=data["close"].size-1
            sheet.cell(row, 387 + 42).value = EMAIndicator(data["close"],5).ema_indicator()[sizzze]
            sheet.cell(row, 388 + 42).value = EMAIndicator(data["close"],10).ema_indicator()[sizzze]
            sheet.cell(row, 389 + 42).value = EMAIndicator(data["close"],20).ema_indicator()[sizzze]
            sheet.cell(row, 390 + 42).value = EMAIndicator(data["close"],50).ema_indicator()[sizzze]
            sheet.cell(row, 391 + 42).value = EMAIndicator(data["close"],100).ema_indicator()[sizzze]
            sheet.cell(row, 392 + 42).value = EMAIndicator(data["close"],200).ema_indicator()[sizzze]
            sheet.cell(row, 387 + 42).value = "EMA10"
            sheet.cell(row, 388 + 42).value = "EMA20"
            sheet.cell(row, 389 + 42).value = "EMA50"
            sheet.cell(row, 390 + 42).value = "EMA100"
            sheet.cell(row, 391 + 42).value = "EMA200"
            sheet.cell(row, 392 + 42).value = "EMA5"
            sheet.cell(row, 393 + 42).value = "EMA5[1]"
            sheet.cell(row, 394 + 42).value = "EMA8"
            sheet.cell(row, 395 + 42).value = "EMA8[1]"
            sheet.cell(row, 396 + 42).value = "EMA13"
            sheet.cell(row, 397 + 42).value = "EMA13[1]"
            sheet.cell(row, 398 + 42).value = "EMA5crossEMA8"
            sheet.cell(row, 399 + 42).value = "EMA8crossEMA13"
            sheet.cell(row, 400 + 42).value = "EMA5crossEMA13"'''
        '''
        if len(dividender.dividends)!=0:
            divideAdder(dividender,data)
        if k == Interval.INTERVAL_1_MONTH:
            ggg=pd.DataFrame(data.iloc[23:])
        else:
            ggg = data
        '''
        data["ll"] = None
        data["hh"] = None
        data["diff"] = None
        data["rdiff"] = None
        data["rdiff_ema"] = None
        data["diff_ema"] = None
        data["avgrel"] = None
        data["avgdiff"] = None
        data["smi"] = None
        data["smi_signal"] = None
        data["ll"] = data["low"].rolling(5, min_periods=1).min()
        data["hh"] = data["high"].rolling(5, min_periods=1).max()
        data["diff"] = data["hh"] - data["ll"]
        data["rdiff"] = data["close"] - (data["hh"] + data["ll"]) / 2
        data["rdiff_ema"] = EMAIndicator(data["rdiff"], 3).ema_indicator()
        data["diff_ema"] = EMAIndicator(data["diff"], 3).ema_indicator()
        data["avgrel"] = EMAIndicator(data["rdiff_ema"], 3).ema_indicator()
        data["avgdiff"] = EMAIndicator(data["diff_ema"], 3).ema_indicator()
        data["smi"] = (data["avgrel"] / (data["avgdiff"] / 2) * 100)
        data["smi_signal"] = EMAIndicator(data["smi"], 3).ema_indicator()
        try:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('smi')).value=data['smi'][len(data['smi']) - 1]
            sheet.cell(row, 2 + kk+7*indicatorlist.index('smi[1]')).value=data['smi'][len(data['smi']) - 2]
            sheet.cell(row, 2 + kk +7*indicatorlist.index('fark')).value=data['smi'][len(data['smi']) - 1]-data['smi'][len(data['smi'])-2]
            sheet.cell(row, 2 + kk +7*indicatorlist.index('smisignal')).value=data['smi_signal'][len(data['smi_signal']) - 1]
            if ((data['smi'][len(data['smi']) - 1] - data['smi_signal'][len(data['smi_signal']) - 1]) * (
                    data['smi'][len(data['smi']) - 2] - data['smi_signal'][len(data['smi_signal']) - 2]) <= 0):
                if ((data['smi'][len(data['smi']) - 1] + data['smi_signal'][len(data['smi_signal']) - 1]) >= (
                        data['smi'][len(data['smi']) - 2] + data['smi_signal'][len(data['smi_signal']) - 2])):
                    sheet.cell(row, 2 + kk +7*indicatorlist.index('smi_smisignal')).value ="crossover"
                else:
                    sheet.cell(row, 2 + kk +7*indicatorlist.index('smi_smisignal')).value ="crossbelow"
            else:
                sheet.cell(row, 2 + kk +7*indicatorlist.index('smi_smisignal')).value =" "
        except:
            var = None
        print("3: " + str(time.time()))
        MACD(data,data['close'])
        sheet.cell(row, 2 + kk+7*indicatorlist.index('lignemacdzerolag')).value = data['MACD'][len(data['MACD'])-1]
        sheet.cell(row, 2 + kk+7*indicatorlist.index('lignesignal')).value = data['Signal_Line'][len(data['Signal_Line'])-1]
        if((data['MACD'][len(data['MACD'])-1]-data['Signal_Line'][len(data['Signal_Line'])-1])*(data['MACD'][len(data['MACD'])-2]-data['Signal_Line'][len(data['Signal_Line'])-2])<=0):
            if((data['MACD'][len(data['MACD'])-1]+data['Signal_Line'][len(data['Signal_Line'])-1])>=(data['MACD'][len(data['MACD'])-2]+data['Signal_Line'][len(data['Signal_Line'])-2])):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_macd')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_macd')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_macd')).value =" "
        print("4: " + str(time.time()))
        data["OBV"] = (np.sign(data["close"].diff().fillna(1)) * data["volume"]).fillna(0).cumsum()
        data["signal"] = ta.wrapper.SMAIndicator(data["OBV"], 21).sma_indicator()
        sheet.cell(row, 2 + kk+7*indicatorlist.index('obv')).value = data['OBV'][len(data['OBV'])-1]
        sheet.cell(row, 2 + kk+7*indicatorlist.index('obv[1]')).value = data['OBV'][len(data['OBV']) - 2]
        sheet.cell(row, 2 + kk+7*indicatorlist.index('farkobv')).value = data['OBV'][len(data['OBV'])-1]-data['OBV'][len(data['OBV']) - 2]
        sheet.cell(row, 2 + kk+7*indicatorlist.index('signal')).value = data['signal'][len(data['signal'])-1]
        if ((data['OBV'][len(data['OBV']) - 1] - data['signal'][len(data['signal']) - 1]) * (
                data['OBV'][len(data['OBV']) - 2] - data['signal'][len(data['signal']) - 2]) <= 0):
            if ((data['OBV'][len(data['OBV']) - 1] + data['signal'][len(data['signal']) - 1]) >= (
                    data['OBV'][len(data['OBV']) - 2] + data['signal'][len(data['signal']) - 2])):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_obv')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_obv')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_obv')).value =" "
        print("5: " + str(time.time()))
        sizzze = data["close"].size - 1
        ema5=EMAIndicator(data['close'],5).ema_indicator()
        ema8 = EMAIndicator(data['close'], 8).ema_indicator()
        ema13 = EMAIndicator(data['close'], 13).ema_indicator()
        sheet.cell(row,2+kk+7*indicatorlist.index('ema5')).value=ema5[sizzze]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema8')).value = ema8[sizzze]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema13')).value = ema13[sizzze]
        sheet.cell(row,2+kk+7*indicatorlist.index('ema5[1]')).value=ema5[sizzze-1]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema8[1]')).value = ema8[sizzze-1]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema13[1]')).value = ema13[sizzze-1]
        rsi1 = ta.wrapper.RSIIndicator(data["close"], 14).rsi()
        kkk = ta.wrapper.StochasticOscillator(rsi1,rsi1,rsi1,14).stoch()
        k = ta.wrapper.SMAIndicator(kkk, 3).sma_indicator()
        d = ta.wrapper.SMAIndicator(k, 3).sma_indicator()
        try:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('k')).value = k[k.size-1]
        except:
            var = None
        sheet.cell(row, 2 + kk+7*indicatorlist.index('d')).value = d[len(d) - 1]
        if ((k[len(k) - 1] - d[len(d) - 1]) * (
                k[len(k) - 2] - d[len(d) - 2]) <= 0):
            if ((k[len(k) - 1] + d[len(d) - 1]) >= (
                    k[len(k) - 2] + d[len(d) - 2])):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_storsi')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_storsi')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('kesişim_storsi')).value =" "
        print("6: " + str(time.time()))
        bollinger_up_rsi, bollinger_down_rsi, basis_rsi = get_bollinger_bands(rsi1)
        bollinger_up_smi, bollinger_down_smi, basis_smi = get_bollinger_bands(data["smi"])
        print("7: " + str(time.time()))
        bolupsmiuse=bollinger_up_smi[bollinger_up_smi.size-1]
        bolupsmiuse2 = bollinger_up_smi[bollinger_up_smi.size - 2]
        boldownsmiuse=bollinger_down_smi[bollinger_down_smi.size-1]
        boldownsmiuse2 = bollinger_down_smi[bollinger_down_smi.size - 2]
        basissmiuse=basis_smi[basis_smi.size-1]
        basissmiuse2 = basis_smi[basis_smi.size - 2]
        smiuse=data['smi'][len(data['smi']) - 1]
        smiuse2 = data['smi'][len(data['smi']) - 2]
        boluprsiuse=bollinger_up_rsi[bollinger_up_rsi.size-1]
        boluprsiuse2 = bollinger_up_rsi[bollinger_up_rsi.size - 2]
        boldownrsiuse=bollinger_down_rsi[bollinger_down_rsi.size-1]
        boldownrsiuse2 = bollinger_down_rsi[bollinger_down_rsi.size - 2]
        basisrsiuse=basis_rsi[basis_rsi.size-1]
        basisrsiuse2 = basis_rsi[basis_rsi.size - 2]
        rsiuse=rsi1[rsi1.size-1]
        rsiuse2 = rsi1[rsi1.size - 2]
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bolupsmi')).value = bolupsmiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('basissmi')).value= basissmiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bollingerdownsmi')).value= boldownsmiuse
        if ((basissmiuse - smiuse) * (
                basissmiuse2 - smiuse2) <= 0):
            if ((basissmiuse + smiuse) >= (
                    basissmiuse2 + smiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basissmidif')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basissmidif')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basissmidif')).value =" "
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bolupbasissmiper')).value = bolupsmiuse-basissmiuse
        if ((bolupsmiuse - smiuse) * (
                bolupsmiuse2 - smiuse2) <= 0):
            if ((bolupsmiuse + smiuse) >= (
                    bolupsmiuse2 + smiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_bolupsmi')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_bolupsmi')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_bolupsmi')).value = " "
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bolupsmidif')).value =bolupsmiuse-smiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bolupsmiper')).value =(bolupsmiuse-smiuse)/smiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boldownsmidif')).value =boldownsmiuse-smiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boldownsmiper')).value =(boldownsmiuse-smiuse)/smiuse
        if ((boldownsmiuse - smiuse) * (
                boldownsmiuse2 - smiuse2) <= 0):
            if ((boldownsmiuse + smiuse) >= (
                    boldownsmiuse2 + smiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownsmi')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownsmi')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownsmi')).value =" "
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boluprsi')).value = boluprsiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('basisrsi')).value = basisrsiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bollingerdownrsi')).value = boldownrsiuse
        if ((basisrsiuse - rsiuse) * (
                basisrsiuse2 - rsiuse2) <= 0):
            if ((basisrsiuse + rsiuse) >= (
                    basisrsiuse2 + rsiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basisrsidif')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basisrsidif')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_basisrsidif')).value =" "
        sheet.cell(row, 2 + kk+7*indicatorlist.index('bolupbasisrsiper')).value = boluprsiuse-basisrsiuse
        if ((boluprsiuse - rsiuse) * (
                boluprsiuse2 - rsiuse2) <= 0):
            if ((boluprsiuse + rsiuse) >= (
                    boluprsiuse2 + rsiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boluprsi')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boluprsi')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boluprsi')).value =" "
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boluprsidif')).value = boluprsiuse-rsiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boluprsiper')).value = (boluprsiuse-rsiuse)/rsiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boldownrsidif')).value = boldownrsiuse-rsiuse
        sheet.cell(row, 2 + kk+7*indicatorlist.index('boldownrsiper')).value = (boldownrsiuse-rsiuse)/rsiuse
        if ((boldownrsiuse - rsiuse) * (
                boldownrsiuse2 - rsiuse2) <= 0):
            if ((boldownrsiuse + rsiuse) >= (
                    boldownrsiuse2 + rsiuse2)):
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownrsi')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownrsi')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk+7*indicatorlist.index('cross_boldownrsi')).value =" "
        print("8: " + str(time.time()))
        ult = ta.wrapper.UltimateOscillator(data["high"],data["low"],data["close"],7,14,28).ultimate_oscillator()
        mfi = ta.wrapper.MFIIndicator(data["high"],data["low"],data["close"],data["volume"],14).money_flow_index()
        stoch = ta.wrapper.StochasticOscillator(data["high"],data["low"],data["close"],14).stoch()
        cci = ta.wrapper.CCIIndicator(data["high"],data["low"],data["close"],14).cci()
        willr=ta.wrapper.WilliamsRIndicator(data["high"],data["low"],data["close"],14).williams_r()
        try:
            momentum = data['close'] / float(data['close'][-15:-14]) * 100
            tke = (ult + mfi + cci + rsi1 + willr + stoch + momentum) / 7
        except:
            tke = (ult + mfi + cci + rsi1 + willr + stoch ) / 6
        # sheet.cell(row, 2 + kk+7*46)
        try:
            mom = data['close'] - data['close'][data["close"].size-11]
            mom2 = data['close'] - data['close'][data["close"].size - 12]
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('mom')).value = mom[mom.size-1]
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('mom[1]')).value = mom2[mom2.size-2]
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('farkmom')).value = mom[mom.size-1] - mom2[mom2.size-2]
        except:
            var = None
        emaline =EMAIndicator(tke,5).ema_indicator()
        tkes=tke[tke.size-1]
        emalines=emaline[emaline.size - 1]
        tkes2=tke[tke.size-2]
        emalines2=emaline[emaline.size - 2]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tke')).value = tkes
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ematke')).value = emalines
        if ((tkes - emalines) * (
                tkes2 - emalines2) <= 0):
            if ((tkes + emalines) >= (
                    tkes2 + emalines2)):
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tkeemacross')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tkeemacross')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tkeemacross')).value =" "
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tke[1]')).value = tkes2
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('tkefark')).value = tkes-tkes2
        sheet.cell(row,2+kk+7*indicatorlist.index('basis-smi')).value=basissmiuse-smiuse
        sheet.cell(row,2+kk+7*indicatorlist.index('basis-rsi')).value=basisrsiuse-rsiuse
        sheet.cell(row,2+kk+7*indicatorlist.index('volume')).value=data["volume"][data["volume"].size-1]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('close')).value = data["close"][data["close"].size - 1]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('rsi')).value = rsiuse
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('rsi[1]')).value = rsiuse2
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('farkrsi')).value = rsiuse - rsiuse2
        cciuse = ta.wrapper.CCIIndicator(data["high"], data["low"], data["close"], 20).cci()
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('cci')).value = cciuse[cciuse.size-1]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('cci[1]')).value = cciuse[cciuse.size-2]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('farkcci')).value = cciuse[cciuse.size-1] - cciuse[cciuse.size-2]
        aroon_up = ta.wrapper.AroonIndicator(data["high"],window=14).aroon_up()
        aroon_down = ta.wrapper.AroonIndicator(data["low"], window=14).aroon_down()
        aroonup1=aroon_up[aroon_up.size-1]
        aroonup2=aroon_up[aroon_up.size-2]
        aroondown1=aroon_down[aroon_up.size-1]
        aroondown2=aroon_down[aroon_up.size-2]
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('aroonup')).value = aroonup1
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('aroonup[1]')).value = aroonup2
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('aroondown')).value = aroondown1
        sheet.cell(row, 2 + kk + 7 * indicatorlist.index('aroondown[1]')).value = aroondown2
        if ((aroonup1 - aroondown1) * (
                aroonup2 - aroondown2) <= 0):
            if ((aroonup1 + aroondown1) >= (
                    aroonup2 + aroondown2)):
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('kesişim_aroon')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('kesişim_aroon')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('kesişim_aroon')).value =" "
        if ((ema5[sizzze] - ema8[sizzze]) * (
                ema5[sizzze-1] - ema8[sizzze-1]) <= 0):
            if ((ema5[sizzze] + ema8[sizzze]) >= (
                    ema5[sizzze-1] + ema8[sizzze-1])):
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema8')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema8')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema8')).value =" "
        if ((ema5[sizzze] - ema13[sizzze]) * (
                ema5[sizzze-1] - ema13[sizzze-1]) <= 0):
            if ((ema5[sizzze] + ema13[sizzze]) >= (
                    ema5[sizzze-1] + ema13[sizzze-1])):
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema13')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema13')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema5_cross_ema13')).value =" "
        if ((ema8[sizzze] - ema13[sizzze]) * (
                ema8[sizzze-1] - ema13[sizzze-1]) <= 0):
            if ((ema8[sizzze] + ema13[sizzze]) >= (
                    ema8[sizzze-1] + ema13[sizzze-1])):
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema8_cross_ema13')).value ="crossover"
            else:
                sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema8_cross_ema13')).value ="crossbelow"
        else:
            sheet.cell(row, 2 + kk + 7 * indicatorlist.index('ema8_cross_ema13')).value =" "
        print("9: " + str(time.time()))
    row+=1
    print("save1: " + str(time.time()))
    wb.save('results.xlsx')
    print("save2: " + str(time.time()))
