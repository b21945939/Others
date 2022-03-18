import time

import openpyxl
import msgpack
import PySimpleGUI as sg
from datetime import datetime

sg.theme("SystemDefaultForReal")
sg.set_options(element_padding=(0, 0), button_element_size=(20, 2), auto_size_buttons=False)
data = []
stockdata = []

def excsave(name):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(0,len(data)):
        for j in range(0,9):
            ws.cell(i+1,j+1).value = data[i][j]
    print(name)
    wb.save(filename = name)
def excadder(event):
    if event == "":
        return
    excel = event
    wb = openpyxl.load_workbook(excel)
    sheet = wb.active
    for row in sheet:
        data2=[]
        for i in range(0,len(row)):
            data2.append(row[i].value)
        while len(data2) < 9:
            data2.append(None)
        data2.append([])
        data.append(data2)
    save(data)


def save(data):
    with open("data.msgpack", "wb") as outfile:
        packed = msgpack.packb(data)
        outfile.write(packed)


def load():
    with open("data.msgpack", "rb") as data_file:
        byte_data = data_file.read()
        data_loaded = msgpack.unpackb(byte_data)
        return data_loaded

def savestock(stock):
    with open("stock.msgpack", "wb") as outfile:
        packed = msgpack.packb(stock)
        outfile.write(packed)


def loadstock():
    with open("stock.msgpack", "rb") as stock_file:
        byte_stock = stock_file.read()
        stock_loaded = msgpack.unpackb(byte_stock)
        return stock_loaded


try:
    data = load()
except:
    save(data)
    data = load()
try:
    stockdata = loadstock()
except:
    savestock(stockdata)
    stockdata = loadstock()
lo = [[sg.Button("Müşteri Ekle"), sg.Input(key='_FILEBROWSE_', enable_events=True, visible=False),
       sg.FileBrowse("Excel Yükle", key='_FILEBROWSE_', file_types=(("Excel", "*.xlsx"),))],
      [sg.Button("Müşteri Çıkar-Düzenle"), sg.Input(key='_FILESAVE_', enable_events=True, visible=False), sg.SaveAs("Excel Kaydet", key='_FILESAVE_',file_types=(("Excel", "*.xlsx"),))],
      [sg.Button("Tümünü Listele"),sg.Button("Stok Girişi")]]
window = sg.Window(title="Müşteri Takip", layout=lo, margins=(50, 50))

def order(a):
    Alist = [[]]
    Alist[0].append(sg.Text("Siparişler:", size=(32,1)))
    b=1
    for i in data[a][9]:
        Alist.append([])
        Alist[b].append(sg.Text(str(b)+".", size=(16, 1)))
        Alist[b].append(sg.Text(i))
        b+=1
    window = sg.Window(title="Sipariş Listesi", margins=(50, 50), modal=True).Layout(
        [Alist])
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
    window.close()

def editor2(a):
    window = sg.Window(title="Düzenleme Paneli",
                       layout=[[sg.Text("Ad Soyad :", size=14), sg.InputText(data[a][0], size=(25, 1))],
                               [sg.Text("Cep Tel :", size=14), sg.InputText(data[a][1], size=(15, 1))],
                               [sg.Text("İş Tel :", size=14), sg.InputText(data[a][2], size=(15, 1))],
                               [sg.Text("E-Posta :", size=14), sg.InputText(data[a][3], size=(40, 1))],
                               [sg.Text("Adres :", size=(14, 1)), sg.Multiline(data[a][4], size=(40, 5))],
                               [sg.Text("Not :", size=(14, 1)), sg.Multiline(data[a][5], size=(40, 5))],
                               [sg.Text("Sosyal Medya :", size=(14)), sg.InputText(data[a][6], size=(25, 1))],
                               [sg.Text("Borç :", size=(14)), sg.InputText(data[a][7], size=(7, 1))],
                               [sg.Text("Tarih ve Saat :", size=14), sg.InputText(data[a][8], size=(17, 1)),
                                sg.Text("", size=20),
                                sg.Button("Siparişler", size=(10, 1)),sg.Button("Kaydet", size=(10, 1)),sg.Button("Sil", size=(5, 1))]], margins=(50, 50), modal=True)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Kaydet":
            for i in range(0, 9):
                data[a][i] = values[i]
            save(data)
        elif event == "Sil":
            data.pop(a)
            save(data)
            break
        elif event == "Siparişler":
            order(a)
    window.close()


def editor():
    window = sg.Window(title="Düzenleme Paneli", layout=[
        [sg.Text("Lütfen düzenlemek istediğiniz ismi giriniz :", size=32), sg.InputText(size=(25, 1)),
         sg.Text("", size=5), sg.Button("Düzenle", size=(10, 1))]], margins=(50, 50), modal=True)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Düzenle":
            keyList=[]
            for i in range(0,len(data)):
                if data[i][0] == values[0]:
                    keyList.append(i)
            if len(keyList)>1:
                miniWindowList(keyList)
            elif len(keyList)==0:
                break
            else:editor2(keyList[0])

    window.close()
def miniWindowList(keyList):
    Alist = [[]]
    Alist[0].append(sg.Text("Aynı İsimden Birden Fazla Bulundu:", size=(32,1)))
    a=1

    for i in keyList:
        Alist.append([])
        Alist[a].append(sg.Text(str(a)+".", size=(16, 1)))
        Alist[a].append(sg.Button("Düzenle",key=str(i)))
        a+=1
    window = sg.Window(title="Müşteri Listesi", margins=(50, 50), modal=True).Layout(
        [Alist])
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif int(event) in keyList:
            editor2(int(event))
            break
    window.close()
def windowList():
    limitDown = 0
    limitUp = 200
    keyList = []
    keyList2 = []
    Mlist = []
    Alist = []
    a = 0
    Alist.append([])
    Alist[a].append(sg.Text("Ad Soyad", size=(16, 1)))
    Alist[a].append(sg.Text("Cep Tel", size=(16, 1)))
    Alist[a].append(sg.Text("İş Tel", size=(16, 1)))
    Alist[a].append(sg.Text("E-Posta", size=(16, 1)))
    Alist[a].append(sg.Text("Adres", size=(16, 1)))
    Alist[a].append(sg.Text("Not", size=(16, 1)))
    Alist[a].append(sg.Text("Sosyal Medya", size=(16, 1)))
    Alist[a].append(sg.Text("Borç", size=(16, 1)))
    Alist[a].append(sg.Text("Tarih ve Saat", size=(16, 1)))
    up = 200
    if (len(data) < 200):
        up = len(data)
    for i in range(0, up):
        Mlist.append([])
        x = ""
        for j in range(0, 9):
            x += str(data[i][j])
            if j in range(4, 6):
                Mlist[a].append(sg.Multiline(x, size=(17, 5)))
            else:
                Mlist[a].append(sg.In(x, size=(17, 1)))
            x = ""
        Mlist[a].append(sg.Button("Tek Kaydet", key=str(a), size=(8, 1)))
        Mlist[a].append(sg.Checkbox("Sil", key=str(a + 200)))
        keyList.append(a)
        keyList2.append(a + 200)
        a += 1
    window = sg.Window(title="Müşteri Listesi", margins=(50, 50), modal=True).Layout(
        [Alist, [sg.Column(Mlist, size=(1400, 600), scrollable=True)],
         [sg.Button("Kaydet"), sg.Button("İşaretlileri Sil"), sg.T(size=45), sg.Button("Geri", disabled=True, size=5),
          sg.T(str(0) + "-" + str(up), key="CHANGE"),
          sg.Button("İleri", size=5) if len(data) > 200 else sg.Button("İleri", size=5, disabled=True)]])
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Kaydet":
            for i in range(0, len(data)):
                for j in range(0, 9):
                    data[i][j] = values[9 * i + j]
            save(data)
        elif event == "İleri":
            window["Geri"].update(disabled=False)
            j = 0
            limitDown = limitUp
            if len(data) > limitUp + 200:
                limitUp += 200
            else:
                limitUp = len(data)
                window["İleri"].update(disabled=True)
                for i in range(limitUp - limitDown, 200):
                    window[str(i)].update(disabled=True)
                for i in range(10 * (limitUp - limitDown), 2000):
                    window[i].update(disabled=True)
            window["CHANGE"].update(str(limitDown) + "-" + str(limitUp))
            for i in range(limitDown, limitUp):
                x = ""
                for k in range(0, 9):
                    x += str(data[i][k])
                    window[j].update(x)
                    x = ""
                    j += 1
        elif event == "Geri":
            if limitUp == len(data):
                for i in range(9 * (limitUp - limitDown), 2000):
                    window[i].update(disabled=False)
                for i in range(limitUp - limitDown, 200):
                    window[str(i)].update(disabled=False)
            window["İleri"].update(disabled=False)
            j = 0
            limitUp = limitDown
            if 200 < limitDown:
                limitDown -= 200
            else:
                limitDown = 0
                window["Geri"].update(disabled=True)
            window["CHANGE"].update(str(limitDown) + "-" + str(limitUp))
            for i in range(limitDown, limitUp):
                x = ""
                for k in range(0, 9):
                    x += str(data[i][k])
                    window[j].update(x)
                    x = ""
                    j += 1
        elif event == "İşaretlileri Sil":
            for i in keyList2:
                if values[str(i)]:
                    data.pop(int(i) - 200 + limitDown)
            save(data)
            break
        elif int(event) in keyList:
            for b in range(0, 9):
                data[int(event) + limitDown][b] = values[9 * (int(event)) + b]
            save(data)
    window.close()


def adder():
    window = sg.Window(title="Ekleme Paneli", layout=[[sg.Text("Ad Soyad :", size=14), sg.InputText(size=(25, 1))],
                                                      [sg.Text("Cep Tel :", size=14), sg.InputText(size=(15, 1))],
                                                      [sg.Text("İş Tel :", size=14), sg.InputText(size=(15, 1))],
                                                      [sg.Text("E-Posta :", size=14), sg.InputText(size=(40, 1))],
                                                      [sg.Text("Adres :", size=(14, 1)), sg.Multiline(size=(40, 5))],
                                                      [sg.Text("Not :", size=(14, 1)), sg.Multiline(size=(40, 5))],
                                                      [sg.Text("Sosyal Medya :", size=14), sg.InputText(size=(25, 1))],
                                                      [sg.Text("Borç :", size=14), sg.InputText(size=(7, 1))],
                                                      [sg.Text("Tarih ve Saat :", size=14),
                                                       sg.InputText(str(datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                                                                    size=(17, 1))],
                                                       [sg.Text("Siparişler :", size=14),sg.Multiline(size=(40, 5)),sg.Text("", size=20),
                                                       sg.Button("Ekle", size=(10, 1))]], margins=(50, 50), modal=True)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Ekle":
            data.append([])
            for i in range(0, 9):
                data[len(data) - 1].append([])
                if values[i]=="":
                    data[len(data) - 1][i] = None
                else:
                    data[len(data) - 1][i] = values[i]
            orders = str(values[9]).split("\n")
            data[len(data) - 1].append([])
            for i in orders:
                data[len(data) - 1][9].append(i)
            save(data)
    window.close()
def stockadder():
    window = sg.Window(title="Stok Ekleme Paneli", layout=[[sg.Text("Ürün :", size=14), sg.InputText(size=(25, 1))],
                                                      [sg.Text("Stok Durumu :", size=14), sg.InputText(size=(15, 1))],
                                                      [sg.Text("Alınan Fiyat :", size=14), sg.InputText(size=(15, 1))],
                                                      [sg.Text("Alınan Yer - Tel :", size=14), sg.Multiline(size=(40, 1))],
                                                      [sg.Text("Tarih ve Saat :", size=14),
                                                       sg.InputText(str(datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                                                                    size=(17, 1)),sg.Button("Ekle", size=(10, 1))]], margins=(50, 50), modal=True)
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Ekle":
            stockdata.append([])
            for i in range(0, 5):
                stockdata[len(stockdata) - 1].append([])
                if(values[i] == ""):
                    stockdata[len(stockdata) - 1][i] = None
                else:
                    stockdata[len(stockdata) - 1][i] = values[i]
            stockdata[len(stockdata) - 1].append([])
            stockdata[len(stockdata) - 1][5] = values[4]
            savestock(stockdata)
    window.close()
def cleanstock(window,keyList,keyList2):
    sList = []
    keyList = []
    keyList2 = []
    a=0
    for i in range(0, len(stockdata)):
        sList.append([])
        for j in range(0, 8):
            sList[i].append([])
        sList[i][0] = sg.In(stockdata[i][0], size=(23))
        sList[i][1] = sg.In(stockdata[i][1], size=23)
        sList[i][2] = sg.In(stockdata[i][2], size=(23))
        sList[i][3] = sg.Multiline(stockdata[i][3], size=(23))
        sList[i][4] = sg.In(stockdata[i][4], size=(23))
        sList[i][5] = sg.In(stockdata[i][5], size=(23))
        sList[i][6] = sg.Button("Sil", key = str(a),size=(5))
        sList[i][7] = sg.Button("Kaydet", key=str(-a-1), size=(5))
        keyList.append(a)
        keyList2.append(-a-1)
        a+=1
    window.close()
    window = sg.Window(title="Stok Paneli", margins=(50, 50), modal=True).Layout(
        [[sg.Button("Ürün Ekle")],
         [sg.T("Ürün Adı", size=(20)), sg.T("Stok Durumu", size=(20)), sg.T("Alınan Fiyat", size=(20)),
          sg.T("Alınan Yer - Tel", size=(20)), sg.T("Düzenleme Tarihi", size=(20)),
          sg.T("Alınma Tarihi", size=(20))],
         [sg.Column(sList, size=(1300, 600), scrollable=True, key="column")]])
    return window,keyList,keyList2
def stock():
    window = sg.Window(title="Boş",modal=False)
    keyList=[]
    keyList2 = []
    window,keyList,keyList2 = cleanstock(window,keyList,keyList2)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == "Ürün Ekle":
            stockadder()
            window,keyList,keyList2 = cleanstock(window,keyList,keyList2)
        elif int(event) in keyList:
            stockdata.pop(int(event))
            savestock(stockdata)
            window,keyList,keyList2 = cleanstock(window, keyList,keyList2)
        elif int(event) in keyList2:
            stockdata[-int(event)-1][0]= values[0 + (-int(event)-1)*6]
            stockdata[-int(event) - 1][1] = values[1 + (-int(event)-1)*6]
            stockdata[-int(event) - 1][2] = values[2 + (-int(event)-1)*6]
            stockdata[-int(event) - 1][3] = values[3 + (-int(event)-1)*6]
            stockdata[-int(event) - 1][4] = str(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            stockdata[-int(event) - 1][5] = values[5 + (-int(event)-1)*6]
            savestock(stockdata)
            window,keyList,keyList2 = cleanstock(window, keyList,keyList2)
    window.close()
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event == "Müşteri Ekle":
        adder()
    elif event == "_FILEBROWSE_":
        excadder(values["_FILEBROWSE_"])
    elif event == "_FILESAVE_":
        excsave(values["_FILESAVE_"])
    elif event == "Müşteri Çıkar-Düzenle":
        editor()
    elif event == "Tümünü Listele":
        windowList()
    elif event == "Stok Girişi":
        stock()
window.close()
